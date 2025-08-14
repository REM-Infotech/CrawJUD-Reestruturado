"""Módulo de geração de arquivos Excel baseados em modelos CrawJUD.

Fornece funcionalidades para criar arquivos XLSX com cabeçalhos e estilos
personalizados, utilizando modelos JSON definidos para o projeto CrawJUD.

"""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING, ClassVar
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl.styles import Font, PatternFill

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

    from crawjud.interface.types import StrPath as StrPath


class MakeTemplates:
    """Implementa geração de arquivos Excel com base em modelos CrawJUD.

    Esta classe permite criar arquivos XLSX utilizando openpyxl, aplicando cabeçalhos
    e estilos personalizados conforme atributos definidos em modelos JSON do CrawJUD.

    Attributes:
        displayname (str): Nome de exibição utilizado no nome do arquivo gerado.
        model_name (str): Nome do modelo JSON utilizado para definir os cabeçalhos.
        temp_dir (str): Diretório temporário para salvar o arquivo XLSX.

    """

    displayname: ClassVar[str] = None
    model_name: ClassVar[str] = None
    temp_dir: ClassVar[str | Path] = None

    def __init__(
        self,
        displayname: str,
        model_name: str,
        temp_dir: str | Path,
    ) -> None:
        """Inicializa a classe MakeTemplates."""
        self.displayname = displayname
        self.model_name = model_name
        self.temp_dir = temp_dir

    def make(self) -> tuple[str, Path]:
        """Create an XLSX file with headers based on the model list.

        Returns:
            tuple: A tuple with the file path and the file name.

        """
        zone_info = ZoneInfo("America/Manaus")
        formated_time = datetime.now(zone_info).strftime("%H-%M-%S")

        dir_file = Path(__file__).parent.resolve().joinpath("models")
        Path(self.temp_dir).mkdir(parents=True, exist_ok=True)

        name_file = f"{self.displayname.upper()} - {formated_time}.xlsx"
        itens_append_file = dir_file.joinpath(f"{self.model_name}.json")
        path_template = str(Path(self.temp_dir).joinpath(name_file))

        # Criar um novo workbook e uma planilha
        workbook = openpyxl.Workbook()
        sheet: Worksheet = workbook.create_sheet("Resultados", 0)

        # Cabeçalhos iniciais
        cabecalhos = ["NUMERO_PROCESSO"]
        list_to_append = []

        if not itens_append_file.exists():
            self.model_name = self.model_name.split("_")[-1]
            itens_append_file = dir_file.joinpath(f"{self.model_name}.json")
            if not itens_append_file.exists():
                self.model_name = "without_model.json"
                itens_append_file = dir_file.joinpath(f"{self.model_name}.json")

        with itens_append_file.open("r") as f:
            itens_append = json.loads(f.read())
            list_to_append.extend(itens_append)

        cabecalhos.extend(list_to_append)
        # Definir estilo
        my_red = openpyxl.styles.colors.Color(rgb="A6A6A6")
        my_fill = PatternFill(patternType="solid", fgColor=my_red)
        bold_font = Font(name="Times New Roman", italic=True)

        # Escrever os cabeçalhos na primeira linha da planilha e aplicar estilo
        for pos, coluna in enumerate(cabecalhos):
            item = sheet.cell(row=1, column=pos + 1, value=coluna.upper())
            item.font = bold_font
            item.fill = my_fill

        # Ajustar a largura das colunas
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                if cell.value:
                    max_length = max(len(str(cell.value)), max_length)
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column].width = adjusted_width

        workbook.save(path_template)
        return path_template, name_file
